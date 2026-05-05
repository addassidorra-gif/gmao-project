from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.exceptions import APIException


def build_xlsx_response(filename, sheet_name, headers, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise APIException("Le module openpyxl est requis pour l'export Excel.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0B2D5B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        worksheet.append([_cell_value(value) for value in row])

    for column in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def build_pdf_response(filename, title, headers, rows):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise APIException("Le module reportlab est requis pour l'export PDF.") from exc

    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    table_data = [headers] + [[str(_cell_value(value)) for value in row] for row in rows]
    if len(table_data) == 1:
        table_data.append(["Aucune donnée"] + [""] * (len(headers) - 1))

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B2D5B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    document.build(story)

    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export_response(file_format, filename, title, headers, rows):
    rows = list(rows)
    if file_format == "xlsx":
        return build_xlsx_response(filename, title[:31], headers, rows)
    if file_format == "pdf":
        return build_pdf_response(filename, title, headers, rows)
    raise APIException("Format d'export non supporté.")


def _cell_value(value):
    if value is None:
        return ""
    if hasattr(value, "tzinfo") and value.tzinfo is not None:
        return timezone.localtime(value).replace(tzinfo=None)
    return value
